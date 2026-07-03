import csv
import re
import uuid
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import TextIOWrapper
from pathlib import Path
from smtplib import SMTPAuthenticationError, SMTPException

from django.core.mail import EmailMessage
from django.conf import settings
from django.db.models import Q
from django.utils import timezone
from openpyxl import load_workbook

from .models import LoanApplication, BankDetail

try:
    from twilio.base.exceptions import TwilioRestException
    from twilio.rest import Client
except ImportError:
    Client = None
    TwilioRestException = Exception




# Add this function right after all import statements
def get_company_config(company_name):
    """
    Return company-specific email and WhatsApp configuration
    """
    from django.conf import settings
    
    if not company_name:
        company_name = 'Mahima'
    
    # Normalize company name to match settings keys
    name_lower = company_name.lower()
    if 'vincit' in name_lower:
        company_key = 'VINCIT'
    elif 'mahima' in name_lower or 'mahim' in name_lower:
        company_key = 'MAHIMA'
    else:
        company_key = company_name.upper().replace(' ', '_')
    
    # Email configuration
    email_config = {
        'host': getattr(settings, f'{company_key}_EMAIL_HOST', settings.EMAIL_HOST),
        'port': getattr(settings, f'{company_key}_EMAIL_PORT', settings.EMAIL_PORT),
        'user': getattr(settings, f'{company_key}_EMAIL_HOST_USER', settings.EMAIL_HOST_USER),
        'password': getattr(settings, f'{company_key}_EMAIL_HOST_PASSWORD', settings.EMAIL_HOST_PASSWORD),
        'from_email': getattr(settings, f'{company_key}_DEFAULT_FROM_EMAIL', settings.DEFAULT_FROM_EMAIL),
        'use_tls': getattr(settings, f'{company_key}_EMAIL_USE_TLS', settings.EMAIL_USE_TLS),
    }
    
    # WhatsApp configuration
    whatsapp_config = {
        'account_sid': getattr(settings, f'{company_key}_TWILIO_ACCOUNT_SID', settings.TWILIO_ACCOUNT_SID),
        'auth_token': getattr(settings, f'{company_key}_TWILIO_AUTH_TOKEN', settings.TWILIO_AUTH_TOKEN),
        'whatsapp_from': getattr(settings, f'{company_key}_TWILIO_WHATSAPP_FROM', settings.TWILIO_WHATSAPP_FROM),
    }
    
    return email_config, whatsapp_config

class NotificationError(Exception):
    pass


class ExcelImportError(Exception):
    pass


def clean_column_name(column: str) -> str:
    """Clean column name by removing quotes, spaces, and special characters"""
    if not column:
        return ""
    column = str(column).strip()
    column = column.strip('"')
    column = column.strip("'")
    column = re.sub(r'\s+', ' ', column)
    return column.strip()


def normalize_column_name(value: str) -> str:
    """Normalize column name for matching"""
    cleaned = clean_column_name(value)
    return cleaned.lower().replace(" ", "_").replace("(", "").replace(")", "").replace("-", "_")


def detect_column(headers, keywords_list):
    """Smart column detection - matches ANY column name"""
    for keywords in keywords_list:
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_clean = str(h).upper().replace(' ', '').replace('_', '').replace('.', '').replace('-', '')
            for kw in keywords:
                kw_clean = kw.upper().replace(' ', '').replace('_', '').replace('.', '').replace('-', '')
                if kw_clean in h_clean or h_clean in kw_clean:
                    return i
    return None


def get_column_mapping(headers):
    """Auto-detect column mapping for ANY Excel file"""
    
    column_patterns = {
        'invoice_no': [
            ['INVOICE', 'INV', 'BILL', 'DOCUMENT', 'DOC NO', 'REFERENCE'],
            ['INVOICE NO', 'INVOICE NUMBER', 'BILL NO', 'BILL NUMBER'],
            ['INVOICE#', 'INV#', 'BILL#']
        ],
        'customer_name': [
            ['PARTY', 'CUSTOMER', 'CLIENT', 'CUST'],
            ['PARTY NAME', 'CUSTOMER NAME', 'CLIENT NAME', 'PARTY DESCRIPTION'],
            ['NAME', 'COMPANY', 'ACCOUNT', 'CUSTOMER NAME']
        ],
        'item_name': [
            ['ITEM NAME', 'ITEM DESCRIPTION', 'PRODUCT', 'MATERIAL'],
            ['ITEM', 'DESCRIPTION', 'PRODUCT NAME']
        ],
        'item_sales_qty': [
            ['QUANTITY', 'QTY', 'SALES QTY', 'ITEM QTY'],
            ['QTY', 'VOLUME', 'SALES QUANTITY']
        ],
        'item_sales_uom': [
            ['UNIT', 'UOM', 'SALES UNIT'],
            ['UOM', 'MEASURE']
        ],
        'item_rate': [
            ['RATE', 'PRICE', 'ITEM RATE', 'SALES RATE'],
            ['RATE', 'UNIT PRICE']
        ],
        'item_amount': [
            ['ITEM AMOUNT', 'PRODUCT AMOUNT', 'MATERIAL VALUE', 'ITEM VALUE'],
            ['ITEM AMT', 'PRODUCT AMT']
        ],
        'employee_name': [
            ['EMPLOYEE NAME', 'RTO', 'SALES PERSON', 'EMPLOYEE'],
            ['EMPLOYEE', 'RTO NAME', 'SALES EXECUTIVE']
        ],
        'invoice_date': [
            ['INVOICE DATE', 'INV DATE', 'DATE'],
            ['BILL DATE', 'DOCUMENT DATE', 'TRANSACTION DATE'],
            ['ISSUE DATE', 'CREATED DATE', 'INVOICE DATE']
        ],
        'invoice_due_date': [
            ['DUE DATE', 'DUE', 'PAYMENT DATE'],
            ['INVOICE DUE DATE', 'DUE ON', 'PAY BY'],
            ['LAST DATE', 'DEADLINE', 'DUE DATE']
        ],
        'invoice_amount': [
            ['BILL AMOUNT', 'INVOICE AMOUNT', 'AMOUNT', 'TOTAL'],
            ['AMT', 'VALUE', 'TOTAL AMOUNT'],
            ['BILL VALUE', 'INVOICE VALUE', 'AMOUNT']
        ],
        'received_amount': [
            ['RECEIVED', 'AMOUNT RECEIVED', 'PAID', 'PAYMENT'],
            ['RECEIVED AMOUNT', 'PAYMENT RECEIVED', 'AMT RECEIVED'],
            ['COLLECTION', 'DEPOSIT', 'RECEIVED']
        ],
        'balance_amount': [
            ['BALANCE', 'OUTSTANDING', 'DUE AMOUNT'],
            ['BALANCE AMOUNT', 'OUTSTANDING AMOUNT', 'REMAINING'],
            ['PENDING', 'UNPAID', 'BALANCE']
        ],
        'overdue_days': [
            ['OVERDUE', 'OVERDUE DAYS', 'DELAY'],
            ['LATE DAYS', 'PAST DUE', 'DAYS']
        ],
        'payment_status': [
            ['PAYMENT STATUS', 'STATUS', 'PAYMENT'],
            ['STATUS', 'PAID STATUS', 'CLEAR STATUS']
        ],
        'credit_term': [
            ['CREDIT TERM', 'RTO', 'CREDIT', 'PAYMENT TERM'],
            ['CREDIT PERIOD', 'TERMS', 'CREDIT TERMS']
        ],
        'whatsapp': [
            ['WHATSAPP', 'MOBILE', 'PHONE', 'CONTACT'],
            ['WHATSAPP NO', 'PHONE NO', 'CELL', 'NUMBER']
        ],
        'email': [
            ['EMAIL', 'MAIL', 'EMAIL ID'],
            ['EMAIL ADDRESS', 'MAIL ID', 'E-MAIL']
        ],
        'comments': [
            ['COMMENT', 'COMMENTS', 'NOTE', 'NOTES', 'REMARK', 'REMARKS'],
            ['INSERT COMMENT', 'COMMENT BOX']
        ]
    }
    
    mapping = {}
    for field, patterns in column_patterns.items():
        idx = detect_column(headers, patterns)
        if idx is not None:
            mapping[field] = headers[idx]
    
    return mapping


def parse_date_smart(value, field_name: str):
    """Parse date from ANY format"""
    if value in (None, ""):
        return None
    
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    
    value_str = str(value).strip()
    
    # Remove time part if present
    if ' ' in value_str and not any(c in value_str for c in ['-', '/']):
        value_str = value_str.split(' ')[0]
    
    # Try all possible date formats
    formats = [
        '%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d.%m.%Y',
        '%d-%b-%Y', '%d %b %Y', '%b %d, %Y', '%d %B %Y',
        '%Y%m%d', '%d%m%Y', '%m%d%Y'
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(value_str, fmt).date()
        except ValueError:
            continue
    
    # Try Excel serial number
    try:
        if isinstance(value, (int, float)):
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=float(value))).date()
    except:
        pass
    
    raise ValueError(f"{field_name} valid date nahi hai: {value_str}")


def parse_amount_smart(value):
    """Parse amount from ANY format"""
    if value in (None, ""):
        return Decimal('0')
    
    try:
        value_str = str(value).strip()
        # Remove currency symbols and commas
        value_str = re.sub(r'[₹$,]', '', value_str)
        value_str = value_str.replace(',', '').strip()
        return Decimal(value_str).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal('0')


def import_loan_applications_from_excel(excel_file, company: str = 'Mahima') -> dict:
    """
    Universal Excel importer - Works with ANY Excel file format
    Automatically detects columns and imports ALL records
    """
    
    # Load Excel file
    extension = Path(getattr(excel_file, "name", "")).suffix.lower()
    
    if extension == ".csv":
        excel_file.seek(0)
        wrapper = TextIOWrapper(excel_file, encoding="utf-8-sig")
        try:
            reader = csv.DictReader(wrapper)
            rows = list(reader)
            headers = reader.fieldnames or []
        finally:
            wrapper.detach()
    elif extension == ".xls":
        raise ExcelImportError("Old `.xls` files are not supported. Please save the file in `.xlsx` format.")
    else:
        try:
            excel_file.seek(0)
            workbook = load_workbook(excel_file, data_only=True)
            worksheet = workbook.active
            
            # Get headers
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(clean_column_name(str(cell.value)))
            
            # Get rows
            rows = []
            for values in worksheet.iter_rows(min_row=2, values_only=True):
                if not values or not any(v not in (None, "") for v in values[:5]):
                    continue
                row = {}
                for idx, header in enumerate(headers):
                    if idx < len(values):
                        row[header] = values[idx]
                rows.append(row)
        except Exception as exc:
            raise ExcelImportError(f"The Excel file could not be read.: {exc}")
    
    if not rows:
        raise ExcelImportError("Excel file khali hai.")
    
    print(f"[Import] Total rows in Excel: {len(rows)}")
    print(f"[Import] Headers found: {headers[:15]}...")
    
    # Auto-detect column mapping
    column_map = get_column_mapping(headers)
    print(f"[Import] Detected mapping: {column_map}")
    
    created_count = 0
    errors = []
    skipped_count = 0
    auto_counter = 1
    today = timezone.localdate()
    last_customer_name = None
    
    for index, row in enumerate(rows):
        try:
            # Get customer name (required)
            customer_name = None
            if 'customer_name' in column_map:
                val = row.get(column_map['customer_name'])
                customer_name = str(val).strip() if val else None
            
            # If we hit "ADVANCE PAYMENTS" section or a similar divider in customer name, stop importing
            if customer_name and ('ADVANCE PAYMENTS' in customer_name.upper() or customer_name.upper().startswith('ADVANCE')):
                print(f"Stopping import at row {index + 2} because of ADVANCE PAYMENTS section")
                break
                
            if not customer_name:
                customer_name = last_customer_name
            else:
                last_customer_name = customer_name
            
            if not customer_name:
                skipped_count += 1
                errors.append(f"Row {index + 2}: Customer Name required - skipping")
                continue
            
            # Parse amounts early so we can check if it is a total/blank row
            invoice_amount = Decimal('0')
            if 'invoice_amount' in column_map:
                invoice_amount = parse_amount_smart(row.get(column_map['invoice_amount']))
                
            # Get invoice number (generate if missing)
            invoice_no = None
            if 'invoice_no' in column_map:
                val = row.get(column_map['invoice_no'])
                invoice_no = str(val).strip() if val else None
            
            if not invoice_no:
                cust_upper = customer_name.upper() if customer_name else ""
                if "TOTAL" in cust_upper or "ADVANCE" in cust_upper or not invoice_amount:
                    continue
                invoice_no = f"AUTO-{company[:3]}-{auto_counter}"
                auto_counter += 1
            
            # Parse dates
            invoice_date = today
            if 'invoice_date' in column_map:
                try:
                    parsed = parse_date_smart(row.get(column_map['invoice_date']), "Invoice Date")
                    if parsed:
                        invoice_date = parsed
                except:
                    pass
            
            invoice_due_date = invoice_date + timedelta(days=30)
            if 'invoice_due_date' in column_map:
                try:
                    parsed = parse_date_smart(row.get(column_map['invoice_due_date']), "Due Date")
                    if parsed:
                        invoice_due_date = parsed
                except:
                    pass
            
            received_amount = Decimal('0')
            if 'received_amount' in column_map:
                received_amount = parse_amount_smart(row.get(column_map['received_amount']))
            
            balance_amount = invoice_amount - received_amount
            if 'balance_amount' in column_map:
                bal = parse_amount_smart(row.get(column_map['balance_amount']))
                if bal > 0:
                    balance_amount = bal
            
            # Parse payment status
            payment_status = 'pending'
            if invoice_amount > 0 and received_amount >= invoice_amount:
                payment_status = 'paid'
            elif invoice_due_date < today and balance_amount > 0:
                payment_status = 'overdue'
            
            if 'payment_status' in column_map:
                status_val = row.get(column_map['payment_status'])
                if status_val:
                    status_str = str(status_val).upper()
                    if any(kw in status_str for kw in ['PAID', 'DONE', 'CLEAR', 'RECEIVED']):
                        payment_status = 'paid'
                    elif any(kw in status_str for kw in ['OVERDUE', 'LATE']):
                        payment_status = 'overdue'
            
            # Parse overdue days
            overdue_days = None
            if invoice_due_date < today and payment_status != 'paid':
                overdue_days = (today - invoice_due_date).days
            
            if 'overdue_days' in column_map:
                val = row.get(column_map['overdue_days'])
                if val and str(val).strip() not in ('', '--', '-', 'N/A', 'nan'):
                    try:
                        overdue_days = int(float(str(val).replace(',', '')))
                    except:
                        pass
            
            # Optional fields
            item_name = ""
            if 'item_name' in column_map:
                val = row.get(column_map.get('item_name', ''))
                item_name = str(val).strip() if val else ""
            
            item_sales_qty = None
            if 'item_sales_qty' in column_map:
                val = row.get(column_map.get('item_sales_qty', ''))
                if val is not None and str(val).strip() not in ('', '--', '-', 'N/A', 'nan'):
                    try:
                        item_sales_qty = Decimal(str(val).replace(',', '').strip())
                    except:
                        pass
                        
            item_sales_uom = ""
            if 'item_sales_uom' in column_map:
                val = row.get(column_map.get('item_sales_uom', ''))
                item_sales_uom = str(val).strip() if val else ""
                
            item_rate = None
            if 'item_rate' in column_map:
                val = row.get(column_map.get('item_rate', ''))
                if val is not None and str(val).strip() not in ('', '--', '-', 'N/A', 'nan'):
                    try:
                        item_rate = Decimal(str(val).replace(',', '').strip())
                    except:
                        pass
                        
            item_amount = None
            if 'item_amount' in column_map:
                val = row.get(column_map.get('item_amount', ''))
                if val is not None and str(val).strip() not in ('', '--', '-', 'N/A', 'nan'):
                    try:
                        item_amount = Decimal(str(val).replace(',', '').strip())
                    except:
                        pass
            
            employee_name = ""
            if 'employee_name' in column_map:
                val = row.get(column_map.get('employee_name', ''))
                employee_name = str(val).strip() if val else ""
            
            credit_term = ""
            if 'credit_term' in column_map:
                val = row.get(column_map.get('credit_term', ''))
                credit_term = str(val).strip() if val else ""
            
            whatsapp = ""
            if 'whatsapp' in column_map:
                val = row.get(column_map.get('whatsapp', ''))
                whatsapp = str(val).strip() if val else ""
            
            email = ""
            if 'email' in column_map:
                val = row.get(column_map.get('email', ''))
                email = str(val).strip() if val else ""
            
            comments = ""
            if 'comments' in column_map:
                val = row.get(column_map.get('comments', ''))
                comments = str(val).strip() if val else ""
            
            # Create record
            record = LoanApplication(
                company=company,
                invoice_no=invoice_no,
                invoice_date=invoice_date,
                customer_name=customer_name,
                item_name=item_name[:255] if item_name else None,
                item_sales_qty=item_sales_qty,
                item_sales_uom=item_sales_uom[:50] if item_sales_uom else None,
                item_rate=item_rate,
                item_amount=item_amount,
                employee_name=employee_name[:255] if employee_name else None,
                credit_term=credit_term[:100] if credit_term else None,
                invoice_due_date=invoice_due_date,
                invoice_amount=invoice_amount,
                received_amount=received_amount,
                balance_amount=balance_amount if balance_amount > 0 else Decimal('0'),
                overdue_days=overdue_days,
                payment_status=payment_status,
                whatsapp=whatsapp[:20] if whatsapp else None,
                email=email[:254] if email else None,
                comments=comments,
                is_active=True
            )
            record.save()
            created_count += 1
            
            if created_count % 50 == 0:
                print(f"[Import] Imported {created_count} records...")
            
        except Exception as exc:
            errors.append(f"Row {index + 2}: {str(exc)}")
            skipped_count += 1
    
    print(f"\n{'='*50}")
    print(f"IMPORT SUMMARY")
    print(f"{'='*50}")
    print(f"Created: {created_count}")
    print(f"Skipped: {skipped_count}")
    print(f"Errors: {len(errors)}")
    print(f"{'='*50}")
    
    return {
        "created_count": created_count,
        "updated_count": 0,
        "errors": errors,
        "skipped_count": skipped_count,
    }


def _get_bank_detail(company_name: str = 'Mahima'):
    """Resolve BankDetail by company name"""
    bank = BankDetail.objects.filter(
        Q(email_sender_name__iexact=company_name) |
        Q(beneficiary_name__iexact=company_name) |
        Q(email_sender_name__icontains=company_name) |
        Q(beneficiary_name__icontains=company_name)
    ).first()
    
    if bank:
        return bank
    
    # Hardcoded fallbacks based on company name
    if company_name and 'vincit' in company_name.lower():
        return type('obj', (object,), {
            'beneficiary_name': 'VINCIT LABS PRIVATE LIMITED',
            'account_no': '50200049027903',
            'bank_name': 'HDFC Bank LTD',
            'branch_name': 'Rohini Sector 9, Delhi.',
            'ifsc_code': 'HDFC0001347',
            'email_sender_name': 'VINCIT LABS PVT LTD'
        })()
    
    # Try any active bank from DB
    bank = BankDetail.objects.filter(is_active=True).first()
    if bank:
        return bank
    
    # Final fallback
    return type('obj', (object,), {
        'beneficiary_name': 'Mahima Life Sciences Pvt Ltd',
        'account_no': 'MLSP96',
        'bank_name': 'HDFC Bank Ltd.',
        'branch_name': 'Rohini Sector 9 – Northex Mall – New Delhi, Delhi',
        'ifsc_code': 'HDFC0001347',
        'email_sender_name': 'Mahima Life Sciences'
    })()


def get_date_format(company_name: str):
    """Return date format based on company"""
    if company_name and 'vincit' in company_name.lower():
        return '%d/%m/%Y'
    return '%d-%m-%Y'


def build_email_message(record: LoanApplication, message: str = None, sender_name: str = None, user=None) -> tuple:
    """Build HTML email with bank details"""
    
    overdue_days = record.current_overdue_days
    
    company_name = sender_name or record.company or 'Mahima'
    is_vincit = 'vincit' in company_name.lower()
    
    if is_vincit:
        company_name = 'Vincit'
    else:
        company_name = 'Mahima'
    
    bank = _get_bank_detail(company_name)
    greeting_company = bank.email_sender_name
    date_format = get_date_format(company_name)
    
    if user:
        final_sender_name = user.get_full_name() or user.username
        final_sender_email = user.email
    else:
        final_sender_name = greeting_company
        final_sender_email = bank.beneficiary_name
    
    if record.received_amount and record.received_amount > 0:
        amount_received_display = f"Rs. {record.received_amount:,.2f}"
        amount_received_cell_style = "padding: 10px; font-size: 13px; color: #15803d; font-weight: bold;"
    else:
        amount_received_display = "--"
        amount_received_cell_style = "padding: 10px; font-size: 13px; color: #64748b;"
    
    if overdue_days > 0:
        overdue_cell_style = "background-color: yellow; color: red; font-weight: bold; text-align: center;"
    else:
        overdue_cell_style = "text-align: center;"
    
    invoice_date_str = record.invoice_date.strftime(date_format)
    due_date_str = record.invoice_due_date.strftime(date_format)
    
    if is_vincit:
        table_headers = """
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">PARTY DESCRIPTION</th>
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">INVOICE NO.</th>
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">INVOICE DATE</th>
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">INVOICE DUE DATE</th>
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">BILL AMOUNT</th>
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">AMOUNT RECEIVED</th>
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">OUTSTANDING AMOUNT</th>
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">OVERDUE DAYS</th>
        """
    else:
        table_headers = """
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">Party Description</th>
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">Invoice No.</th>
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">Invoice Date</th>
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">Invoice Due Date</th>
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">Bill Amount</th>
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">Amount Received</th>
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">Outstanding Amount</th>
            <th style="text-align: left; padding: 10px; font-size: 13px; font-weight: bold; text-transform: uppercase;">Overdue Days</th>
        """
    
    if message:
        main_content = f"<p>{message.replace('\n', '<br>')}</p>"
    else:
        main_content = f"""
        <p>Dear Sir/Madam,</p>
        <p>Greetings from {greeting_company}</p>
        <p>We hope you are doing well.</p>
        <p>We would like to kindly inform you that the payment for the below-mentioned invoice is now due. We would appreciate it if you could please share an update on the expected payment date.</p>
        """
    
    html_message = f"""
    <html>
    <body style="font-family: Arial, sans-serif;">
        {main_content}
        <br>
        <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; border: 1px solid #cbd5e1; font-family: Arial, sans-serif;">
            <thead>
                <tr style="background-color: #e8eaf6; color: #1e3a8a;">
                    {table_headers}
                </tr>
            </thead>
            <tbody>
                <tr>
                    <td style="padding: 10px; font-size: 13px;">{record.customer_name}</td>
                    <td style="padding: 10px; font-size: 13px;">{record.invoice_no}</td>
                    <td style="padding: 10px; font-size: 13px;">{invoice_date_str}</td>
                    <td style="padding: 10px; font-size: 13px; font-weight: bold;">{due_date_str}</td>
                    <td style="padding: 10px; font-size: 13px;">Rs. {record.invoice_amount:,.2f}</td>
                    <td style="{amount_received_cell_style}">{amount_received_display}</td>
                    <td style="padding: 10px; font-size: 13px;">Rs. {record.balance_amount:,.2f}</td>
                    <td style="padding: 10px; font-size: 13px; {overdue_cell_style}">{overdue_days:02d}</td>
                </tr>
            </tbody>
        </table>
        <br>
        <br>
        <p><span style="text-decoration: underline; color: red; font-weight: bold;">Our Virtual Account Bank Detail are Given Below –</span></p>
        <br>
        <ul style="list-style-type: disc; padding-left: 20px; font-weight: bold; color: #1e3a8a;">
            <li style="margin-bottom: 5px;">Beneficiary Name &nbsp;&nbsp;&nbsp;: {bank.beneficiary_name}</li>
            <li style="margin-bottom: 5px;">Beneficiary A/C. No. : {bank.account_no}</li>
            <li style="margin-bottom: 5px;">Bank Name &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {bank.bank_name}</li>
            <li style="margin-bottom: 5px;">Branch Name &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {bank.branch_name}</li>
            <li style="margin-bottom: 5px;">IFSC Code &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {bank.ifsc_code}</li>
        </ul>
        <br>
        <p>Best Regards,<br>
        {final_sender_name}<br>
        <small>{final_sender_email}</small></p>
    </body>
    </html>
    """
    
    if message:
        text_content = message
    else:
        text_content = f"""Dear Sir/Madam,

Greetings from {greeting_company}
We hope you are doing well.
We would like to kindly inform you that the payment for the below-mentioned invoice is now due. We would appreciate it if you could please share an update on the expected payment date."""
    
    text_message = f"""{text_content}

Party Description: {record.customer_name}
Invoice No: {record.invoice_no}
Invoice Date: {invoice_date_str}
Invoice Due Date: {due_date_str}
Bill Amount: Rs. {record.invoice_amount:,.2f}
Amount Received: {amount_received_display}
Outstanding Amount: Rs. {record.balance_amount:,.2f}
Overdue Days: {overdue_days}

Our Virtual Account Bank Detail are Given Below –

Beneficiary Name: {bank.beneficiary_name}
Beneficiary A/C. No.: {bank.account_no}
Bank Name: {bank.bank_name}
Branch Name: {bank.branch_name}
IFSC Code: {bank.ifsc_code}

Best Regards,
{final_sender_name}
({final_sender_email})"""
    
    return html_message, text_message


def build_whatsapp_message(record: LoanApplication, message: str = None, sender_name: str = None, user=None) -> str:
    """Build WhatsApp message"""
    
    overdue_days = record.current_overdue_days
    
    company_name = sender_name or record.company or 'Mahima'
    is_vincit = 'vincit' in company_name.lower()
    
    if is_vincit:
        company_name = 'Vincit'
    else:
        company_name = 'Mahima'
    
    bank = _get_bank_detail(company_name)
    greeting_company = bank.email_sender_name
    date_format = get_date_format(company_name)
    
    invoice_date_str = record.invoice_date.strftime(date_format)
    due_date_str = record.invoice_due_date.strftime(date_format)
    
    if record.received_amount and record.received_amount > 0:
        amount_received_display = f"Rs. {record.received_amount:,.2f}"
    else:
        amount_received_display = "--"
        
    if message:
        intro = message
    else:
        intro = f"""Greetings from {greeting_company}. We hope you are doing well.
We would like to kindly inform you that the payment for the below-mentioned invoice is now due. We would appreciate it if you could please share an update on the expected payment date."""
    
    message_body = f"""*{greeting_company} - Payment Reminder*

{intro}

*Party:* {record.customer_name}
*Invoice No:* {record.invoice_no}
*Invoice Date:* {invoice_date_str}
*Due Date:* {due_date_str}
*Bill Amount:* Rs. {record.invoice_amount:,.2f}
*Amount Received:* {amount_received_display}
*Outstanding:* Rs. {record.balance_amount:,.2f}
*Overdue:* {overdue_days} days

*Our Virtual Account Bank Detail are Given Below -*
• *Beneficiary Name:* {bank.beneficiary_name}
• *Beneficiary A/C. No.:* {bank.account_no}
• *Bank Name:* {bank.bank_name}
• *Branch Name:* {bank.branch_name}
• *IFSC Code:* {bank.ifsc_code}

Please arrange payment at earliest.

Thank you."""
    
    return message_body


# def send_email_notification(record: LoanApplication, message: str = None, sender_name: str = None, user=None) -> str:
#     if not settings.EMAIL_HOST_USER or not settings.EMAIL_HOST_PASSWORD:
#         return "email credentials missing"
    
#     if not record.email:
#         return "no email address"
    
#     try:
#         bank = _get_bank_detail(sender_name or record.company)
#         final_company_name = sender_name or bank.email_sender_name
        
#         html_message, text_message = build_email_message(record, message=message, sender_name=sender_name, user=user)
        
#         # Generate unique Message-ID for this email
#         domain = settings.DEFAULT_FROM_EMAIL.split('@')[-1] if '@' in settings.DEFAULT_FROM_EMAIL else 'localhost'
#         new_message_id = f"<{uuid.uuid4()}@{domain}>"
        
#         email = EmailMessage(
#             subject=f"Payment Reminder - Invoice {record.invoice_no}",
#             body=html_message,
#             from_email=f"{final_company_name} <{settings.DEFAULT_FROM_EMAIL}>",
#             to=[record.email],
#             reply_to=[user.email if user and user.email else settings.DEFAULT_FROM_EMAIL]
#         )
#         email.content_subtype = "html"
        
#         # Email Threading: If previous email was sent, add threading headers
#         # This makes follow-up emails appear in the SAME conversation thread
#         # in Gmail, Outlook, etc. instead of as separate emails
#         email.extra_headers = {'Message-ID': new_message_id}
        
#         if record.email_message_id:
#             # In-Reply-To: points to the LAST email in the thread
#             email.extra_headers['In-Reply-To'] = record.email_message_id
#             # References: contains the full chain of Message-IDs
#             email.extra_headers['References'] = record.email_message_id
        
#         email.send()
        
#     except SMTPAuthenticationError as exc:
#         raise NotificationError("Email login failed. Please use a Gmail App Password.") from exc
#     except SMTPException as exc:
#         raise NotificationError(f"Email send nahi hua: {exc}") from exc
    
#     # Save the Message-ID for future threading + update timestamp
#     record.email_message_id = new_message_id
#     record.last_email_sent_at = timezone.now()
#     record.save(update_fields=["last_email_sent_at", "email_message_id", "updated_at"])
    
#     from .models import FollowUp
#     FollowUp.objects.create(
#         application=record,
#         followup_type='Email',
#         followup_date=timezone.localdate(),
#         status='Completed',
#         notes=f"System email reminder sent to {record.email}"
#     )
#     return "sent"
def send_email_notification(record: LoanApplication, message: str = None, sender_name: str = None, sender_company: str = None, user=None) -> str:
    # Determine which company to use
    company = sender_company
    if sender_name and not company:
        if 'vincit' in sender_name.lower():
            company = 'Vincit'
        elif 'mahima' in sender_name.lower() or 'mahim' in sender_name.lower():
            company = 'Mahima'
    if not company:
        company = record.company or 'Mahima'
    
    # Get company-specific configuration
    email_config, _ = get_company_config(company)
    
    # Check if credentials exist
    if not email_config['user'] or not email_config['password']:
        return f"email credentials missing for {company}"
    
    if not record.email:
        return "no email address"
    
    try:
        bank = _get_bank_detail(company)
        final_company_name = sender_name or bank.email_sender_name
        
        html_message, text_message = build_email_message(record, message=message, sender_name=sender_name, user=user)
        
        domain = email_config['from_email'].split('@')[-1] if '@' in email_config['from_email'] else 'localhost'
        new_message_id = f"<{uuid.uuid4()}@{domain}>"
        
        # Create email connection with company-specific settings
        from django.core.mail import get_connection
        connection = get_connection(
            host=email_config['host'],
            port=email_config['port'],
            username=email_config['user'],
            password=email_config['password'],
            use_tls=email_config['use_tls'],
        )
        
        email = EmailMessage(
            subject=f"Payment Reminder - Invoice {record.invoice_no}",
            body=html_message,
            from_email=email_config['from_email'],
            to=[record.email],
            reply_to=[user.email if user and user.email else email_config['user']],
            connection=connection
        )
        email.content_subtype = "html"
        email.extra_headers = {'Message-ID': new_message_id}
        
        if record.email_message_id:
            email.extra_headers['In-Reply-To'] = record.email_message_id
            email.extra_headers['References'] = record.email_message_id
        
        email.send()
        
    except SMTPAuthenticationError as exc:
        raise NotificationError(f"Email login failed for {company}: {str(exc)}") from exc
    except SMTPException as exc:
        raise NotificationError(f"Email send failed for {company}: {str(exc)}") from exc
    
    # Update record
    record.email_message_id = new_message_id
    record.last_email_sent_at = timezone.now()
    record.save(update_fields=["last_email_sent_at", "email_message_id", "updated_at"])
    
    # Create follow-up record
    from .models import FollowUp
    FollowUp.objects.create(
        application=record,
        followup_type='Email',
        followup_date=timezone.localdate(),
        status='Completed',
        notes=f"Email reminder sent to {record.email} using {company} credentials"
    )
    return "sent"

# def send_whatsapp_notification(record: LoanApplication, message: str = None, sender_name: str = None, user=None) -> str:
#     if not settings.TWILIO_ACCOUNT_SID or not settings.TWILIO_AUTH_TOKEN:
#         return "twilio credentials missing"
#     if Client is None:
#         return "twilio package missing"
    
#     if not record.whatsapp:
#         return "no whatsapp number"
    
#     message = build_whatsapp_message(record, message=message, sender_name=sender_name, user=user)
    
#     try:
#         client = Client(settings.TWILIO_ACCOUNT_SID, settings.TWILIO_AUTH_TOKEN)
#         client.messages.create(
#             body=message,
#             from_=settings.TWILIO_WHATSAPP_FROM,
#             to=f"whatsapp:{record.whatsapp}",
#         )
#     except TwilioRestException as exc:
#         raise NotificationError(f"WhatsApp message send nahi hua: {exc}") from exc
    
#     record.last_whatsapp_sent_at = timezone.now()
#     record.save(update_fields=["last_whatsapp_sent_at", "updated_at"])
    
#     from .models import FollowUp
#     FollowUp.objects.create(
#         application=record,
#         followup_type='WhatsApp',
#         followup_date=timezone.localdate(),
#         status='Completed',
#         notes=f"System WhatsApp reminder sent to {record.whatsapp}"
#     )
#     return "sent"
def send_whatsapp_notification(record: LoanApplication, message: str = None, sender_name: str = None, sender_company: str = None, user=None) -> str:
    # Determine which company to use
    company = sender_company
    if sender_name and not company:
        if 'vincit' in sender_name.lower():
            company = 'Vincit'
        elif 'mahima' in sender_name.lower() or 'mahim' in sender_name.lower():
            company = 'Mahima'
    if not company:
        company = record.company or 'Mahima'
    
    # Get company-specific configuration
    _, whatsapp_config = get_company_config(company)
    
    # Check if credentials exist
    if not whatsapp_config['account_sid'] or not whatsapp_config['auth_token']:
        return f"twilio credentials missing for {company}"
    if Client is None:
        return "twilio package missing"
    
    if not record.whatsapp:
        return "no whatsapp number"
    
    message_body = build_whatsapp_message(record, message=message, sender_name=sender_name, user=user)
    
    try:
        client = Client(whatsapp_config['account_sid'], whatsapp_config['auth_token'])
        client.messages.create(
            body=message_body,
            from_=whatsapp_config['whatsapp_from'],
            to=f"whatsapp:{record.whatsapp}",
        )
    except TwilioRestException as exc:
        raise NotificationError(f"WhatsApp message send failed for {company}: {str(exc)}") from exc
    
    # Update record
    record.last_whatsapp_sent_at = timezone.now()
    record.save(update_fields=["last_whatsapp_sent_at", "updated_at"])
    
    # Create follow-up record
    from .models import FollowUp
    FollowUp.objects.create(
        application=record,
        followup_type='WhatsApp',
        followup_date=timezone.localdate(),
        status='Completed',
        notes=f"WhatsApp reminder sent to {record.whatsapp} using {company} number"
    )
    return "sent"

# def send_record_notifications(record: LoanApplication, message: str = None, sender_name: str = None, user=None) -> dict:
#     results = {
#         "email": send_email_notification(record, message=message, sender_name=sender_name, user=user),
#         "whatsapp": send_whatsapp_notification(record, message=message, sender_name=sender_name, user=user),
#     }
#     return results

def send_record_notifications(record: LoanApplication, message: str = None, sender_name: str = None, sender_company: str = None, user=None) -> dict:
    company = sender_company
    if sender_name and not company:
        if 'vincit' in sender_name.lower():
            company = 'Vincit'
        elif 'mahima' in sender_name.lower() or 'mahim' in sender_name.lower():
            company = 'Mahima'
    if not company:
        company = record.company or 'Mahima'
        
    results = {
        "email": send_email_notification(record, message=message, sender_name=sender_name, sender_company=company, user=user),
        "whatsapp": send_whatsapp_notification(record, message=message, sender_name=sender_name, sender_company=company, user=user),
    }
    return results
def build_message(record: LoanApplication, sender_name: str = None, user=None) -> str:
    return build_whatsapp_message(record, sender_name=sender_name, user=user)